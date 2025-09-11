# app.py (Version Final - script corrigé)
import os
import json
import io
import hashlib
from datetime import datetime, date
from typing import List, Dict, Any, Optional

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
#================================================================================================
# ---------- Helpers : normalisation dates & sanitization ----------
from datetime import datetime, date

def _to_date_obj(val):
    """
    Retourne un datetime.date à partir de:
      - datetime.date (retourné tel quel)
      - datetime.datetime (on prend .date())
      - str (formats courants : 'YYYY-MM-DD' ou 'YYYY/MM/DD')
      - None / autre -> date.today()
    """
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(val, fmt).date()
            except Exception:
                pass
    return date.today()

def _sanitize_row_for_storage(row: dict) -> dict:
    """
    Transforme toute valeur date/datetime en string 'YYYY-MM-DD' pour éviter
    erreurs de sérialisation JSON lors de l'écriture en fichier / base.
    """
    sanitized = {}
    for k, v in row.items():
        if isinstance(v, (datetime, date)):
            sanitized[k] = v.strftime("%Y-%m-%d")
        else:
            sanitized[k] = v
    return sanitized

#================================================================================================

# ---------------------------
# Configuration & thème
# ---------------------------
st.set_page_config(page_title="Work Order Management", layout="wide")

# Palette bleu / gris pastel - CSS simple
st.markdown(
    """
    <style>
    :root {
        --primary-1: #2b6ea3; /* bleu principal */
        --primary-2: #6ea0c8; /* bleu clair */
        --muted: #6b7280;
        --card-bg: #f5f8fb;
        --input-bg: #eef4fb;
    }
    .app-header { padding:10px 12px; border-radius:8px; color: white; margin-bottom:10px; }
    .card { background: var(--card-bg); padding:12px; border-radius:10px; box-shadow: 0 6px 14px rgba(0,0,0,0.04); margin-bottom:12px; }
    .small-muted { color: var(--muted); font-size:12px; }
    .stButton>button { background: linear-gradient(90deg, var(--primary-1), var(--primary-2)); color: white; }
    .sidebar .stButton>button { background: var(--primary-1); color: white; }
    .stTextInput>div>input, .stSelectbox>div>div>div, .stDateInput>div>input { background: var(--input-bg); }
    </style>
    """,
    unsafe_allow_html=True
)

# Titre principal
st.title("Work Order Management")

# ---------------------------
# Data files
# ---------------------------
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
os.makedirs(DATA_DIR, exist_ok=True)

FILES = {
    "bon_travail": os.path.join(DATA_DIR, "bon_travail.json"),
    "liste_pdr": os.path.join(DATA_DIR, "liste_pdr.json"),
    "users": os.path.join(DATA_DIR, "users.json"),
    "options_description_probleme": os.path.join(DATA_DIR, "options_description_probleme.json"),
    "options_poste_de_charge": os.path.join(DATA_DIR, "options_poste_de_charge.json"),
}

# Initial values (copiées depuis ta Version Final)
INITIAL_DESCRIPTIONS = [
    'P.M.I.01-Panne au niveau du capos',"P.M.I.02-problème d'éjecteur de moule",'P.M.I.03-Blocage  moule',
    'P.M.I.04-Problème de tiroir','P.M.I.05-Cassure vis sortie plaque carotte','P.M.I.06-Blocage de la plaque carotte',
    'P.M.I.07-Vis de noyaux endommagé','P.M.I.08-Problème noyau',"P.M.I.09-Problème vis d'injection",'P.M.I.10-Réducteur',
    'P.M.I.11-Roue dentée ','P.M.I.12-PB grenouillère','P.M.I.13-Vis de pied endommagé','P.M.I.14-Colonnes de guidage ',
    "P.M.I.15-Fuite matiére au niveau de la buse d'injection",
    'P.E.I.01-PB capteur ','P.E.I.02-PB galet (fin de course)','P.E.I.03-PB moteur électrique','P.E.I.04-Capteur linéaire',
    'P.E.I.05-Armoire électrique ','P.E.I.06-Écran/tactile',"P.E.I.07-Machine s'allume pas","P.E.I.08-PB d'électrovanne",
    'P.E.I.09-PB connecteur ','P.E.I.10-Système magnétique',
    'P.H.I.01-PB flexible','P.H.I.02-PB raccord','P.H.I.03-PB vérin','P.H.I.04-PB distributeur','P.H.I.05-PB pompe',
    'P.H.I.06-PB filtre','P.H.I.07-PB au niveau huile','P.H.I.08-PB fuite huile','P.H.I.09-PB préchauffage',
    'P.H.I.10-PB lubrification du canalisation de grenouillère',
    'P.P.I.01-PB de pression','P.P.I.02-Remplissage matière ','P.P.I.03-Alimentation matiére ',
    'P.P.I.04-Flexible pneumatique','P.P.I.05-PB raccord',
    'P.T.I.01-PB collier chauffante','P.T.I.02-PB de thermocouple','P.T.I.03-Zone de chauffage en arrêt',
    'P.T.I.04-PB refroidisseur',"P.T.I.05-PB pression d'eau",'P.T.I.06-PB température sécheur',
    'P.T.I.07-Variation de la température (trop élever/trop bas )'
]

INITIAL_POSTES = [
    'ASL011','ASL021','ASL031','ASL041','ASL051','ASL061','ASL071',
    'ASL012','ASL022','ASL032','ASL042','ASL052','ASL062','ASL072',
    'ACL011','ACL021','ACL031','ACL041','ACL051','ACL061','ACL071','APCL011','APCL021','APCL031',
    'CL350-01 HOUSING','CL350-02 HOUSING','CL350-03 BRAKET ','CL120-01 SUR MOULAGE (LEVIET)','CL120-02 SUR MOULAGE (LEVIET)',
    'M. Shifter Ball', 'M. Knob clip-lever MA','M. Knob clip-lever MB6', 'M. Guides for trigger', 'M. Damper',
    'M. MB6-HIGH HOUSING', 'M. MB6-LOW HOUSING','M. MA-HIGH HOUSING', 'M. MA-LOW HOUSING', 'M. BRAKET MA'
]

# ---------------------------
# Fichiers utilitaires atomiques
# ---------------------------
def atomic_write(path: str, obj: Any) -> None:
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def load_json(path: str) -> Any:
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def ensure_data_files():
    if load_json(FILES["bon_travail"]) is None:
        atomic_write(FILES["bon_travail"], [])
    if load_json(FILES["liste_pdr"]) is None:
        atomic_write(FILES["liste_pdr"], [])
    if load_json(FILES["users"]) is None:
        atomic_write(FILES["users"], [])
    if load_json(FILES["options_description_probleme"]) is None:
        atomic_write(FILES["options_description_probleme"], INITIAL_DESCRIPTIONS.copy())
    if load_json(FILES["options_poste_de_charge"]) is None:
        atomic_write(FILES["options_poste_de_charge"], INITIAL_POSTES.copy())

ensure_data_files()

def read_options(name: str) -> List[str]:
    path = FILES.get(name)
    if path and os.path.exists(path):
        return load_json(path) or []
    return []

def write_options(name: str, opts: List[str]):
    path = FILES.get(name)
    if path:
        atomic_write(path, opts)


# ---------------------------
# Hash mot de passe
# ---------------------------
def hash_password(pwd: str) -> str:
    return hashlib.sha256((pwd or "").encode("utf-8")).hexdigest()

# ---------------------------
# CRUD Bons (colonnes originales)
# ---------------------------
BON_COLUMNS = [
    "code","date","arret_declare_par","poste_de_charge","heure_declaration","machine_arreter",
    "heure_debut_intervention","heure_fin_intervention","technicien","description_probleme",
    "action","pdr_utilisee","observation","resultat","condition_acceptation","dpt_maintenance","dpt_qualite","dpt_production"
]

def read_bons() -> List[Dict[str, Any]]:
    arr = load_json(FILES["bon_travail"])
    return arr or []

def write_bons(arr: List[Dict[str, Any]]):
    atomic_write(FILES["bon_travail"], arr)

def get_bon_by_code(code: str) -> Optional[Dict[str, Any]]:
    for r in read_bons():
        if str(r.get("code","")) == str(code):
            return r
    return None

def add_bon(bon: Dict[str, Any]) -> None:
    bons = read_bons()
    if get_bon_by_code(bon.get("code")) is not None:
        raise ValueError("Code déjà présent")

    # --- Normaliser avant stockage ---
    entry = {k: bon.get(k, "") for k in BON_COLUMNS}
    # Si une valeur est un objet date/datetime → convertir en string
    for k, v in entry.items():
        if isinstance(v, (datetime, date)):
            entry[k] = v.strftime("%Y-%m-%d")

    bons.append(entry)
    write_bons(bons)

    # décrémenter PDR si fourni (si PDR existe)
    pdr_code = str(entry.get("pdr_utilisee","")).strip()
    if pdr_code:
        pdrs = read_pdr()
        for i, p in enumerate(pdrs):
            if str(p.get("code","")).strip() == pdr_code:
                q = int(p.get("quantite",0) or 0)
                p["quantite"] = max(0, q-1)
                pdrs[i] = p
                write_pdr(pdrs)
                break

def update_bon(code: str, updates: Dict[str, Any]) -> None:
    bons = read_bons()
    found = False
    for i, r in enumerate(bons):
        if str(r.get("code","")) == str(code):
            for k in BON_COLUMNS:
                if k in updates:
                    val = updates[k]
                    # Normaliser les dates avant stockage
                    if isinstance(val, (datetime, date)):
                        val = val.strftime("%Y-%m-%d")
                    r[k] = val
            bons[i] = r
            found = True
            break
    if not found:
        raise KeyError("Code introuvable")
    write_bons(bons)

def compute_progress(bon: Dict[str, Any]) -> int:
    """
    Retourne un pourcentage d'avancement (int 0..100) basé sur les colonnes:
      dpt_production, dpt_maintenance, dpt_qualite (logique inchangée mais sécurisée).
    - 100% si les 3 dpt == "Valider"
    - Sinon, % = (nombre de champs non vides parmi BON_COLUMNS / len(BON_COLUMNS)) * 100
    Toujours renvoie un int entre 0 et 100.
    """
    try:
        # Si les 3 validations sont 'Valider' => 100%
        if bon.get("dpt_production") == "Valider" and \
           bon.get("dpt_maintenance") == "Valider" and \
           bon.get("dpt_qualite") == "Valider":
            return 100

        # Comptage sécurisé des champs non vides
        filled = sum(1 for k, v in bon.items() if v not in ("", None))
        total_columns = len(BON_COLUMNS) if len(BON_COLUMNS) > 0 else 1
        progress = int((filled / total_columns) * 100)

        # Borner la valeur entre 0 et 100
        return max(0, min(100, progress))
    except Exception:
        # En cas d'erreur imprévue, retourner 0 (sécurité)
        return 0

def delete_bon(code: str) -> None:
    bons = read_bons()
    bons = [r for r in bons if str(r.get("code","")) != str(code)]
    write_bons(bons)

# ---------------------------
# PDR CRUD (garde les fonctions si tu veux la page)
# ---------------------------
PDR_COLUMNS = ["code","remplacement","nom_composant","quantite"]

def read_pdr() -> List[Dict[str, Any]]:
    arr = load_json(FILES["liste_pdr"])
    return arr or []

def write_pdr(arr: List[Dict[str, Any]]):
    atomic_write(FILES["liste_pdr"], arr)

def upsert_pdr(rec: Dict[str, Any]):
    pdrs = read_pdr()
    code = str(rec.get("code","")).strip()
    if not code:
        raise ValueError("Code PDR requis")
    for i,p in enumerate(pdrs):
        if str(p.get("code","")).strip() == code:
            pdrs[i] = {"code": code, "remplacement": rec.get("remplacement",""), "nom_composant": rec.get("nom_composant",""), "quantite": int(rec.get("quantite",0))}
            write_pdr(pdrs)
            return
    pdrs.append({"code": code, "remplacement": rec.get("remplacement",""), "nom_composant": rec.get("nom_composant",""), "quantite": int(rec.get("quantite",0))})
    write_pdr(pdrs)

def delete_pdr_by_code(code: str):
    pdrs = read_pdr()
    pdrs = [p for p in pdrs if str(p.get("code","")).strip() != str(code).strip()]
    write_pdr(pdrs)

# ---------------------------
# Users helpers
# ---------------------------
def read_users() -> List[Dict[str, Any]]:
    arr = load_json(FILES["users"])
    return arr or []

def write_users(arr: List[Dict[str, Any]]):
    atomic_write(FILES["users"], arr)

def get_user(username: str) -> Optional[Dict[str,Any]]:
    for u in read_users():
        if u.get("username","") == username:
            return u
    return None

def create_user(username: str, password: str, role: str):
    users = read_users()
    if get_user(username):
        raise ValueError("Utilisateur existe déjà")
    new_id = (len(users) + 1) if users else 1
    users.append({"id": new_id, "username": username, "password_hash": hash_password(password), "role": role})
    write_users(users)



# ---------------------------
# plot_pareto (défini avant usage pour les periode)
# ---------------------------
def plot_pareto(df: pd.DataFrame, period: str = "day", top_n_labels: int = 3):
    s = pd.to_datetime(df['date'], errors='coerce').dropna()
    if s.empty:
        st.info("Aucune date valide pour tracer le Pareto.")
        return
    if period == "day":
        groups = s.dt.strftime("%Y-%m-%d")
        xlabel = "Jour"
    elif period == "week":
        groups = s.dt.strftime("%Y-W%U")
        xlabel = "Semaine"
    else:
        groups = s.dt.strftime("%Y-%m")
        xlabel = "Mois"

    counts = groups.value_counts().sort_values(ascending=False)
    total = counts.sum()
    if total == 0:
        st.info("Pas assez de données.")
        return
    cum_pct = 100 * counts.cumsum() / total

    fig, ax1 = plt.subplots(figsize=(10,4))
    fig.patch.set_facecolor("#f8fbff")
    ax1.set_facecolor("#ffffff")

    x = np.arange(len(counts))
    cmap = plt.get_cmap("viridis")
    colors = cmap(np.linspace(0.2, 0.8, len(counts)))
    bars = ax1.bar(x, counts.values, color=colors, edgecolor="#2b2b2b", linewidth=0.2)
    ax1.set_xticks(x)
    ax1.set_xticklabels(counts.index.tolist(), rotation=45, ha='right', fontsize=9)
    ax1.set_ylabel("Nombre d'interventions")
    ax1.set_xlabel(xlabel)
    ax1.set_title(f"Pareto ({period}) - total = {total}", fontsize=12, weight="bold")
    ax1.grid(axis="y", alpha=0.12)

    ax2 = ax1.twinx()
    ax2.plot(x, cum_pct.values, color='#ff7f0e', marker='o', linewidth=2)
    ax2.set_ylim(0, 110)
    ax2.set_ylabel("Pourcentage cumulé (%)", color='#ff7f0e')
    ax2.tick_params(axis='y', labelcolor='#ff7f0e')
    ax2.axhline(80, color='grey', linestyle='--', alpha=0.6)

    # annotate top
    top = counts.head(top_n_labels)
    for idx, (label, val) in enumerate(counts.items()):
        if idx < top_n_labels:
            pct = val/total*100
            ax1.text(idx, val + max(counts.values)*0.02, f"{val} ({pct:.1f}%)", ha='center', fontsize=9, bbox=dict(boxstyle="round", alpha=0.18))

    plt.tight_layout()
    st.pyplot(fig)

    st.markdown("**Périodes les plus impactées :**")
    for i, (label, val) in enumerate(top.items(), start=1):
        st.write(f"{i}. **{label}** — {val} interventions — {val/total*100:.1f}%")


# ---------------------------
# plot_paretoo (par type de problème avec filtre)
# ---------------------------
def plot_paretoo(df: pd.DataFrame, top_n_labels: int = 5):
    if "description_probleme" not in df.columns:
        st.warning("La colonne 'description_probleme' est absente des données.")
        return

    # --- Sélecteur de type de problème ---
    types = df["description_probleme"].dropna().unique().tolist()
    types = sorted([str(t) for t in types])
    selected_type = st.selectbox("Filtrer par type de problème :", ["Tous"] + types)

    # --- Filtrage ---
    if selected_type != "Tous":
        df = df[df["description_probleme"] == selected_type]

    counts = df["description_probleme"].value_counts().sort_values(ascending=False)
    total = counts.sum()
    if total == 0:
        st.info("Pas assez de données après filtrage.")
        return

    cum_pct = 100 * counts.cumsum() / total

    fig, ax1 = plt.subplots(figsize=(10,4))
    fig.patch.set_facecolor("#f8fbff")
    ax1.set_facecolor("#ffffff")

    x = np.arange(len(counts))
    cmap = plt.get_cmap("viridis")
    colors = cmap(np.linspace(0.2, 0.8, len(counts)))
    ax1.bar(x, counts.values, color=colors, edgecolor="#2b2b2b", linewidth=0.2)

    ax1.set_xticks(x)
    ax1.set_xticklabels(counts.index.tolist(), rotation=45, ha='right', fontsize=9)
    ax1.set_ylabel("Nombre d'occurrences")
    ax1.set_xlabel("Type de problème")
    ax1.set_title(f"Pareto des problèmes - total = {total}", fontsize=12, weight="bold")
    ax1.grid(axis="y", alpha=0.12)

    ax2 = ax1.twinx()
    ax2.plot(x, cum_pct.values, color='#ff7f0e', marker='o', linewidth=2)
    ax2.set_ylim(0, 110)
    ax2.set_ylabel("Pourcentage cumulé (%)", color='#ff7f0e')
    ax2.tick_params(axis='y', labelcolor='#ff7f0e')
    ax2.axhline(80, color='grey', linestyle='--', alpha=0.6)

    # --- Annoter les top problèmes ---
    top = counts.head(top_n_labels)
    for idx, (label, val) in enumerate(counts.items()):
        if idx < top_n_labels:
            pct = val/total*100
            ax1.text(idx, val + max(counts.values)*0.02,
                     f"{val} ({pct:.1f}%)",
                     ha='center', fontsize=9,
                     bbox=dict(boxstyle="round", alpha=0.18))

    plt.tight_layout()
    st.pyplot(fig)

    st.markdown("**Problèmes les plus récurrents :**")
    for i, (label, val) in enumerate(top.items(), start=1):
        st.write(f"{i}. **{label}** — {val} fois — {val/total*100:.1f}%")



# ---------------------------
# Export Excel utilitaire (page d'export possible)
# ---------------------------
def export_excel(bons: List[Dict[str,Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bon de travail"
    start_row = 1
    for col_idx, h in enumerate(BON_COLUMNS, start=1):
        ws.cell(row=start_row, column=col_idx).value = h
        ws.cell(row=start_row, column=col_idx).font = Font(bold=True)
    rownum = start_row + 1
    for r in bons:
        for col_idx, h in enumerate(BON_COLUMNS, start=1):
            ws.cell(row=rownum, column=col_idx).value = r.get(h, "")
        rownum += 1
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ---------------------------
# Helpers session: charger / clear valeur formulaire (page-scoped)
# ---------------------------
def load_bon_into_session(bon: Dict[str, Any], page_name: str):
    """Charge un bon dans st.session_state en préfixant par page_name."""
    updates = {}
    for k in BON_COLUMNS:
        val = bon.get(k, "")
        if k == "date" and val:
            try:
                if isinstance(val, str):
                    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y"):
                        try:
                            val = datetime.strptime(val, fmt).date()
                            break
                        except Exception:
                            continue
                elif isinstance(val, datetime):
                    val = val.date()
            except Exception:
                val = date.today()
        updates[f"{page_name}_form_{k}"] = val
    st.session_state.update(updates)



def clear_form_session(page_name: str):
    for k in BON_COLUMNS:
        st.session_state[f"{page_name}_form_{k}"] = ""

# ---------------------------
# Header - logo si disponible
# ---------------------------
logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo REGAL-PNG.png")
if os.path.exists(logo_path):
    try:
        st.image(logo_path, width=200)
    except Exception:
        pass

# ---------------------------
# Session state utilisateur et manager vérif persistante
# ---------------------------
if "user" not in st.session_state:
    st.session_state.user = None
if "role" not in st.session_state:
    st.session_state.role = None
if "manager_verified" not in st.session_state:
    st.session_state.manager_verified = False

# ---------------------------
# Sidebar: Login & création utilisateur (avec gestion manager_verified)
# ---------------------------
st.sidebar.title("Connexion")
users = read_users()

if st.session_state.user:
    st.sidebar.success(f"Connecté: {st.session_state.user} ({st.session_state.role})")
    if st.sidebar.button("Se déconnecter", key="btn_logout"):
        st.session_state.user = None
        st.session_state.role = None
        st.rerun()
else:
    # Formulaire de connexion
    login_user = st.sidebar.text_input("Nom d'utilisateur", key="login_user")
    login_pwd = st.sidebar.text_input("Mot de passe", key="login_pwd", type="password")
    if st.sidebar.button("Se connecter", key="btn_login"):
        u = get_user(login_user)
        if not u or u.get("password_hash") != hash_password(login_pwd):
            st.sidebar.error("Identifiants invalides.")
        else:
            st.session_state.user = u["username"]
            st.session_state.role = u["role"]
            st.sidebar.success(f"Bienvenue {u['username']} ({u['role']})")
            st.rerun()

# Création de compte (validation manager required)
st.sidebar.markdown("---")
st.sidebar.subheader("Créer un compte (nécessite vérification manager)")

if not st.session_state.manager_verified:
    mgr_name = st.sidebar.text_input("Manager (username)", key="mgr_name")
    mgr_pwd = st.sidebar.text_input("Manager (mdp)", key="mgr_pwd", type="password")
    if st.sidebar.button("Vérifier manager", key="btn_check_mgr"):
        mgr = get_user(mgr_name)
        if not mgr or mgr.get("password_hash") != hash_password(mgr_pwd) or mgr.get("role") != "manager":
            st.sidebar.error("Vérification échouée.")
        else:
            st.sidebar.success("Manager vérifié — complétez la création.")
            st.session_state.manager_verified = True
            st.rerun()
else:
    # Champs pour créer un nouvel utilisateur (après vérification)
    new_user = st.sidebar.text_input("Nouveau utilisateur", key="new_user")
    new_pwd = st.sidebar.text_input("Nouveau mdp", key="new_pwd", type="password")
    new_role = st.sidebar.selectbox("Rôle", ["production","maintenance","qualite","manager"], key="new_role")
    if st.sidebar.button("Créer utilisateur", key="btn_create_user"):
        try:
            create_user(new_user.strip(), new_pwd, new_role)
            st.sidebar.success("Utilisateur créé.")
            st.session_state.manager_verified = False
        except Exception as e:
            st.sidebar.error(str(e))
    if st.sidebar.button("Annuler", key="btn_cancel_create"):
        st.session_state.manager_verified = False
        st.rerun()

# Si aucun utilisateur, proposer création manager initial (one-shot)
if not users:
    st.warning("Aucun utilisateur trouvé — créez un manager initial.")
    with st.form("init_mgr_form"):
        mgru = st.text_input("Manager username", value="manager", key="init_mgr_user")
        mgrp = st.text_input("Manager password", type="password", key="init_mgr_pwd")
        submitted = st.form_submit_button("Créer manager initial")   # <-- OBLIGATOIRE
    if submitted:
        if not mgru or not mgrp:
            st.error("Remplissez les champs.")
        else:
            create_user(mgru.strip(), mgrp, "manager")
            st.success("Manager initial créé — connectez-vous.")
            st.rerun()


# ---------------------------
# Sidebar menu (Pages)
# ---------------------------
menu = st.sidebar.radio("Pages", ["Dashboard","Production","Maintenance","Qualité"])

# ---------------------------
# Permissions helper
# ---------------------------
def allowed(page: str) -> bool:
    role = st.session_state.role
    if role == "manager":
        return True
    if role == "production" and page == "Production":
        return True
    if role == "maintenance" and page == "Maintenance":
        return True
    if role == "qualite" and page == "Qualité":
        return True
    return False

def page_dashboard():
    st.markdown(
        '<div class="app-header" style="background: linear-gradient(90deg, #2b6ea3, #6ea0c8);">'
        "<h3 style='margin:6px 0'>Tableau de bord — Pareto & résumé</h3></div>",
        unsafe_allow_html=True,
    )

    bons = read_bons()
    if not bons:
        st.info("Aucun bon enregistré.")
        return

    df = pd.DataFrame(bons)

    # Colonnes affichées dans le tableau résumé
    display_cols = ["code", "date", "dpt_production", "dpt_maintenance", "dpt_qualite"]

    c1, c2 = st.columns([3, 1])
    # Choix du Top N
    topn = c2.number_input("Top N", min_value=1, max_value=10, value=3, key="dash_topn")

    # ---------------------------
    # Analyse Pareto
    # ---------------------------
    st.markdown("### Analyse Pareto")
    col_p1, col_p2 = st.columns(2)

    with col_p1:
        st.markdown("**Pareto par période**")
        # Sélecteur de période
        period = st.selectbox("Filtrer par période :", ["day", "week", "month"], key="pareto_period")
        try:
            plot_pareto(df, period=period, top_n_labels=topn)
        except Exception as e:
            st.warning(f"Erreur dans plot_pareto : {e}")

    with col_p2:
        st.markdown("**Pareto par type de problème**")
        try:
            plot_paretoo(df, top_n_labels=topn)
        except Exception as e:
            st.warning(f"Erreur dans plot_paretoo : {e}")

    # ---------------------------
    # Aperçu des bons
    # ---------------------------
    st.markdown("---")
    st.subheader("Aperçu (derniers bons d’intervention)")
    st.dataframe(df.sort_values(by="date", ascending=False)[display_cols], height=320)

    # ---------------------------
    # État d’avancement coloré
    # ---------------------------
    # Calcul progression
    df["Progression (%)"] = df.apply(compute_progress, axis=1)

    # Palette de 8 couleurs distinctes (progression par tranche)
    palette = [
        "#e74c3c",  # 0-12% rouge
        "#e67e22",  # 13-25% orange
        "#f39c12",  # 26-37% jaune foncé
        "#f1c40f",  # 38-50% jaune
        "#2ecc71",  # 51-62% vert clair
        "#27ae60",  # 63-75% vert foncé
        "#3498db",  # 76-87% bleu
        "#2c3e50",  # 88-100% gris/bleu foncé
    ]

    def color_row(row):
        val = row["Progression (%)"]
        # découpage en 8 tranches
        idx = min(val // 13, 7)  # 0-12, 13-25, ..., 88-100
        color = palette[idx]
        return [f"background-color: {color}; color: white;" for _ in row]

    st.markdown("### État d'avancement des bons")
    st.progress(int(df["Progression (%)"].mean()))  # moyenne globale

    styled_df = (
        df[["code", "date", "dpt_production", "dpt_maintenance", "dpt_qualite", "Progression (%)"]]
        .sort_values(by="date", ascending=False)
        .style.apply(color_row, axis=1)
    )
    st.dataframe(styled_df, height=300)


# ---------------------------
# Page: Bons (Production / Maintenance / Qualité)
# ---------------------------
def page_bons(page_name: str):
    # Si un chargement est en attente (après un submit), l'appliquer maintenant
    if "pending_load" in st.session_state:
        row, pg = st.session_state.pop("pending_load")
        load_bon_into_session(row, pg)
    
    st.markdown(f'<div class="app-header" style="background: linear-gradient(90deg, #2b6ea3, #6ea0c8);"><h3 style="margin:6px 0">{page_name} — Gestion des bons</h3></div>', unsafe_allow_html=True)
    if not allowed(page_name):
        st.warning("Vous n'avez pas la permission pour cette page.")
        return

    bons = read_bons()
    df = pd.DataFrame(bons) if bons else pd.DataFrame(columns=BON_COLUMNS)
    codes = df["code"].astype(str).tolist() if not df.empty else []

    # Charger / Nouveau
    st.subheader("Charger / Nouveau")
    col_load1, col_load2 = st.columns([3,1])
    sel_key = f"sel_{page_name}"
    sel_code = col_load1.selectbox("Charger un bon existant (optionnel)", options=[""] + codes, key=sel_key)
    if col_load2.button("Charger", key=f"btn_load_{page_name}") and sel_code:
        bon = get_bon_by_code(sel_code)
        if bon:
            load_bon_into_session(bon, page_name)
            st.rerun()
    if col_load2.button("Nouveau", key=f"btn_new_{page_name}"):
        clear_form_session(page_name)
        st.rerun()



    # Définition des champs éditables par fenêtre
    production_allowed = {
        "code","date", "heure_declaration", "description_probleme", "arret_declare_par",
        "poste_de_charge", "machine_arreter", "resultat", "condition_acceptation", "dpt_production"
    }
    maintenance_allowed = {
        "heure_debut_intervention", "heure_fin_intervention", "technicien", "observation", "dpt_maintenance"
    }
    qualite_allowed = {
        "heure_debut_intervention", "heure_fin_intervention", "technicien", "observation", "dpt_qualite"
    }

    if page_name.lower().startswith("production"):
        editable_set = production_allowed
    elif page_name.lower().startswith("maintenance"):
        editable_set = maintenance_allowed
    elif page_name.lower().startswith("qualit") or page_name.lower().startswith("qualité"):
        editable_set = qualite_allowed
    else:
        editable_set = set()

    # Initialiser session keys pour ce page si manquants
    for k in BON_COLUMNS:
        sk = f"{page_name}_form_{k}"
        if sk not in st.session_state:
            st.session_state[sk] = ""

    # Formulaire unique (id unique par page)
    form_id = f"form_bon_{page_name}"
    
    with st.form(form_id,):
        c1, c2, c3 = st.columns([2,1,1])

        # Code
        code_key = f"{page_name}_form_code"
        code = c1.text_input("Code", value=st.session_state.get(code_key, ""), key=code_key, disabled=("code" not in editable_set))

        # Date (string -> date)
        # --- Normaliser la valeur de session avant d'instancier le widget (corrige JSON error) ---
        date_key = f"{page_name}_form_date"

        # Récupérer la valeur stockée (peut être string, date, datetime, None)
        sess_val = st.session_state.get(date_key, None)
        safe_date = _to_date_obj(sess_val)          # convertit proprement en datetime.date
        # Forcer la session à contenir un datetime.date (streamlit widget serialise correctement)
        st.session_state[date_key] = safe_date

        # Maintenant créer le widget en passant une datetime.date sûr
        date_input = c1.date_input("Date", value=st.session_state[date_key], key=date_key, disabled=("date" not in editable_set))



        # Arrêt déclaré par
        arret_key = f"{page_name}_form_arret_declare_par"
        arret = c1.text_input("Arrêt déclaré par", value=st.session_state.get(arret_key, ""), key=arret_key, disabled=("arret_declare_par" not in editable_set))

        # Poste de charge
        poste_key = f"{page_name}_form_poste_de_charge"
        postes = read_options("options_poste_de_charge")
        poste_default = st.session_state.get(poste_key, "")

        if "poste_de_charge" in editable_set:
            # construire options dynamiques
            poste_options = [""] + postes + ["Autres..."]
            index_default = poste_options.index(poste_default) if poste_default in poste_options else 0
            poste = c2.selectbox("Poste de charge", poste_options, index=index_default, key=poste_key)

            if poste == "Autres...":
                new_poste_key = f"{page_name}_new_poste"
                add_poste_confirm_key = f"{page_name}_add_poste_confirm"
                new_poste = c2.text_input("Ajouter nouveau poste (libre)", key=new_poste_key)

                # case à cocher pour confirmer l'ajout immédiat (permis dans un form)
                add_now = c2.checkbox("Enregistrer ce poste maintenant", key=add_poste_confirm_key)

                if new_poste and add_now:
                    new_poste_clean = new_poste.strip()
                    if new_poste_clean:
                        # relire options au cas où elles ont changé
                        postes_current = read_options("options_poste_de_charge")
                        if new_poste_clean not in postes_current:
                            postes_current.append(new_poste_clean)
                            write_options("options_poste_de_charge", postes_current)
                            # Mettre à jour la session pour conserver la sélection et forcer re-render
                            st.session_state[poste_key] = new_poste_clean
                            # vider le champ d'ajout et décocher
                            st.session_state[new_poste_key] = ""
                            st.session_state[add_poste_confirm_key] = False
                            st.experimental_rerun()
                        else:
                            st.info("Ce poste existe déjà.")
        else:
            # affichage en lecture seule
            poste_options = [""] + postes
            index_default = poste_options.index(poste_default) if poste_default in poste_options else 0
            c2.selectbox("Poste de charge", poste_options, index=index_default, disabled=True, key=f"{poste_key}_ro")
            poste = poste_default


        # Heure déclaration
        heure_key = f"{page_name}_form_heure_declaration"
        heure_declaration = c2.text_input("Heure de déclaration", value=st.session_state.get(heure_key, ""), key=heure_key, disabled=("heure_declaration" not in editable_set))

        # Machine arrêtée?
        machine_key = f"{page_name}_form_machine_arreter"
        machine = c2.selectbox("Machine arrêtée?", ["","Oui","Non"], index=(["","Oui","Non"].index(st.session_state.get(machine_key,"")) if st.session_state.get(machine_key,"") in ["","Oui","Non"] else 0), key=machine_key, disabled=("machine_arreter" not in editable_set))

        # Heures intervention
        debut_key = f"{page_name}_form_heure_debut_intervention"
        fin_key = f"{page_name}_form_heure_fin_intervention"
        debut = c3.text_input("Heure début", value=st.session_state.get(debut_key, ""), key=debut_key, disabled=("heure_debut_intervention" not in editable_set))
        fin = c3.text_input("Heure fin", value=st.session_state.get(fin_key, ""), key=fin_key, disabled=("heure_fin_intervention" not in editable_set))

        # Technicien
        tech_key = f"{page_name}_form_technicien"
        technicien = c3.text_input("Technicien", value=st.session_state.get(tech_key, ""), key=tech_key, disabled=("technicien" not in editable_set))

        # Description problème
        # Description problème (Production uniquement avec "Autres...")
        desc_key = f"{page_name}_form_description_probleme"
        descs = read_options("options_description_probleme")
        desc_default = st.session_state.get(desc_key, "")

        if "description_probleme" in editable_set:
            description = st.selectbox("Description", [""] + descs + ["Autres..."], 
                                    index=([""]+descs+["Autres..."]).index(desc_default) 
                                    if desc_default in ([""]+descs+["Autres..."]) else 0, 
                                    key=desc_key)
            if st.session_state.get(desc_key) == "Autres...":
                new_desc = st.text_input("Ajouter nouvelle description", key=f"{page_name}_new_desc")
                if new_desc:
                    new_desc_clean = new_desc.strip()
                    if new_desc_clean and new_desc_clean not in descs:
                        descs.append(new_desc_clean)
                        write_options("options_description_probleme", descs)
                        st.session_state[desc_key] = new_desc_clean
                        description = new_desc_clean
        else:
            _idx = ([""]+descs).index(desc_default) if desc_default in ([""]+descs) else 0
            st.selectbox("Description", [""] + descs, index=_idx, disabled=True, key=f"{desc_key}_ro")
            description = desc_default


        # Action
        action_key = f"{page_name}_form_action"
        action = st.text_input("Action", value=st.session_state.get(action_key, ""), key=action_key, disabled=("action" not in editable_set))

        # PDR utilisée
        pdr_key = f"{page_name}_form_pdr_utilisee"
        pdr_used = st.text_input("PDR utilisée (code)", value=st.session_state.get(pdr_key, ""), key=pdr_key, disabled=("pdr_utilisee" not in editable_set))

        # Observation
        obs_key = f"{page_name}_form_observation"
        observation = st.text_input("Observation", value=st.session_state.get(obs_key, ""), key=obs_key, disabled=("observation" not in editable_set))

        # Résultat
        res_key = f"{page_name}_form_resultat"
        resultat = st.selectbox("Résultat", ["","Accepter","Refuser","Accepter avec condition"], index=(["","Accepter","Refuser","Accepter avec condition"].index(st.session_state.get(res_key,"")) if st.session_state.get(res_key,"") in ["","Accepter","Refuser","Accepter avec condition"] else 0), key=res_key, disabled=("resultat" not in editable_set))

        # Condition d'acceptation
        cond_key = f"{page_name}_form_condition_acceptation"
        cond = st.text_input("Condition d'acceptation", value=st.session_state.get(cond_key, ""), key=cond_key, disabled=("condition_acceptation" not in editable_set))

        # Dpts
        dpt_m_key = f"{page_name}_form_dpt_maintenance"
        dpt_q_key = f"{page_name}_form_dpt_qualite"
        dpt_p_key = f"{page_name}_form_dpt_production"
        dpt_m = st.selectbox("Dpt Maintenance", ["","Valider","Non Valider"], index=(["","Valider","Non Valider"].index(st.session_state.get(dpt_m_key,"")) if st.session_state.get(dpt_m_key,"") in ["","Valider","Non Valider"] else 0), key=dpt_m_key, disabled=("dpt_maintenance" not in editable_set))
        dpt_q = st.selectbox("Dpt Qualité", ["","Valider","Non Valider"], index=(["","Valider","Non Valider"].index(st.session_state.get(dpt_q_key,"")) if st.session_state.get(dpt_q_key,"") in ["","Valider","Non Valider"] else 0), key=dpt_q_key, disabled=("dpt_qualite" not in editable_set))
        dpt_p = st.selectbox("Dpt Production", ["","Valider","Non Valider"], index=(["","Valider","Non Valider"].index(st.session_state.get(dpt_p_key,"")) if st.session_state.get(dpt_p_key,"") in ["","Valider","Non Valider"] else 0), key=dpt_p_key, disabled=("dpt_production" not in editable_set))

        submit_key = f"submit_{page_name}"
        submitted = st.form_submit_button("Ajouter / Mettre à jour", key=f"submit_{page_name}")


        if submitted:
            code_v = st.session_state.get(code_key, "").strip()
            date_val = st.session_state.get(date_key)
            if isinstance(date_val, (datetime, date)):
                date_v = date_val.strftime("%Y-%m-%d")
            else:
                date_v = str(date_val) if date_val else date.today().strftime("%Y-%m-%d")

            row = {k: "" for k in BON_COLUMNS}
            row.update({
                "code": code_v,
                "date": date_v,
                "arret_declare_par": st.session_state.get(arret_key, ""),
                "poste_de_charge": st.session_state.get(poste_key, ""),
                "heure_declaration": st.session_state.get(heure_key, ""),
                "machine_arreter": st.session_state.get(machine_key, ""),
                "heure_debut_intervention": st.session_state.get(debut_key, ""),
                "heure_fin_intervention": st.session_state.get(fin_key, ""),
                "technicien": st.session_state.get(tech_key, ""),
                "description_probleme": st.session_state.get(desc_key, ""),
                "action": st.session_state.get(action_key, ""),
                "pdr_utilisee": st.session_state.get(pdr_key, ""),
                "observation": st.session_state.get(obs_key, ""),
                "resultat": st.session_state.get(res_key, ""),
                "condition_acceptation": st.session_state.get(cond_key, ""),
                "dpt_maintenance": st.session_state.get(dpt_m_key, ""),
                "dpt_qualite": st.session_state.get(dpt_q_key, ""),
                "dpt_production": st.session_state.get(dpt_p_key, "")
            })
            if submitted:
                try:
                    if code_v == "":
                        st.error("Le champ Code est requis pour ajouter ou mettre à jour un bon.")
                    else:
                        if any(c.get("code","") == code_v for c in read_bons()):
                            update_bon(code_v, row)
                            st.success("Bon mis à jour.")
                        else:
                            add_bon(row)
                            st.success("Bon ajouté.")
                    # Bouton pour supprimer le bon actuellement chargé
                    if code and get_bon_by_code(code):
                        if st.form_submit_button("🗑️ Supprimer ce bon", key=f"del_btn_{page_name}"):
                            delete_bon(code)
                            st.success(f"Bon {code} supprimé avec succès.")
                            clear_form_session(page_name)
                            st.rerun()

                    # Décaler le chargement à l’exécution suivante
                    st.session_state["pending_load"] = (row, page_name)
                    st.rerun()
                except Exception as e:
                    st.error(str(e))

    # ---- Gestion des options (EN DEHORS du form) ----
    if page_name.lower().startswith("production"):
        st.markdown("### Gérer les postes et descriptions")

        col_g1, col_g2 = st.columns(2)

        # ---- Postes ----
        new_poste_out = col_g1.text_input("Poste", key=f"{page_name}_add_poste_out")

        c1_add, c1_del = col_g1.columns([1,1])
        # Ajouter
        if c1_add.button("Ajouter poste", key=f"{page_name}_btn_add_poste_out"):
            if new_poste_out:
                postes_current = read_options("options_poste_de_charge")
                if new_poste_out.strip() not in postes_current:
                    postes_current.append(new_poste_out.strip())
                    write_options("options_poste_de_charge", postes_current)
                    st.success(f"Poste '{new_poste_out}' ajouté.")
                    st.rerun()
                else:
                    st.info("Ce poste existe déjà.")
        # Supprimer
        if c1_del.button("Supprimer poste", key=f"{page_name}_btn_del_poste_out"):
            postes_current = read_options("options_poste_de_charge")
            if new_poste_out.strip() in postes_current:
                postes_current.remove(new_poste_out.strip())
                write_options("options_poste_de_charge", postes_current)
                st.warning(f"Poste '{new_poste_out}' supprimé.")
                st.rerun()
            else:
                st.info("Ce poste n'existe pas.")

        # ---- Descriptions ----
        new_desc_out = col_g2.text_input("Description", key=f"{page_name}_add_desc_out")
        c2_add, c2_del = col_g2.columns([1,1])
        # Ajouter
        if c2_add.button("Ajouter description", key=f"{page_name}_btn_add_desc_out"):
            if new_desc_out:
                descs_current = read_options("options_description_probleme")
                if new_desc_out.strip() not in descs_current:
                    descs_current.append(new_desc_out.strip())
                    write_options("options_description_probleme", descs_current)
                    st.success(f"Description '{new_desc_out}' ajoutée.")
                    st.rerun()
                else:
                    st.info("Cette description existe déjà.")
        # Supprimer
        if c2_del.button("Supprimer description", key=f"{page_name}_btn_del_desc_out"):
            descs_current = read_options("options_description_probleme")
            if new_desc_out.strip() in descs_current:
                descs_current.remove(new_desc_out.strip())
                write_options("options_description_probleme", descs_current)
                st.warning(f"Description '{new_desc_out}' supprimée.")
                st.rerun()
            else:
                st.info("Cette description n'existe pas.")


    
    st.markdown('</div>', unsafe_allow_html=True)

    # Recherche & Liste (unique)
    st.markdown("---")
    st.subheader("Recherche & Liste")
    search_by_key = f"{page_name}_search_by"
    term_key = f"{page_name}_term"
    search_by = st.selectbox("Rechercher par", ["Code","Date","Poste de charge","Dpt"], key=search_by_key)
    term = st.text_input("Terme de recherche", key=term_key)
    if st.button("Rechercher", key=f"btn_search_{page_name}"):
        res = []
        for r in read_bons():
            col = ""
            if search_by == "Code":
                col = r.get("code","")
            elif search_by == "Date":
                col = r.get("date","")
            elif search_by == "Poste de charge":
                col = r.get("poste_de_charge","")
            else:
                col = r.get("dpt_production","") + r.get("dpt_maintenance","") + r.get("dpt_qualite","")
            if term.lower() in str(col).lower():
                res.append(r)
        if not res:
            st.info("Aucun enregistrement trouvé.")
        else:
            st.dataframe(pd.DataFrame(res), height=250)

    # Tous les bons
    st.subheader("Tous les bons")
    all_df = pd.DataFrame(read_bons())
    if not all_df.empty:
        st.dataframe(all_df.sort_values(by="date", ascending=False), height=300)
        sel_key = f"{page_name}_sel_code"
        sel = st.selectbox("Sélectionner un code", options=[""] + all_df["code"].astype(str).tolist(), key=sel_key)
        if sel:
            if st.button("Afficher JSON", key=f"showjson_{page_name}"):
                st.json(get_bon_by_code(sel))
            if st.button("Supprimer", key=f"del_{page_name}"):
                delete_bon(sel)
                st.success("Supprimé")
                st.rerun()
    else:
        st.info("Aucun bon à afficher.")

# ---------------------------
# Page PDR (optionnelle)
# ---------------------------
def page_pdr():
    st.header("Pièces - PDR (liste_pdr)")
    pdrs = read_pdr()
    df = pd.DataFrame(pdrs) if pdrs else pd.DataFrame(columns=PDR_COLUMNS)
    st.dataframe(df, height=250)
    with st.form("form_pdr"):
        code = st.text_input("Code PDR", key="pdr_code")
        remplacement = st.text_input("Remplacement", key="pdr_remp")
        nom = st.text_input("Nom composant", key="pdr_nom")
        quantite = st.number_input("Quantité", min_value=0, value=0, key="pdr_qte")
        if st.form_submit_button("Enregistrer PDR"):
            try:
                upsert_pdr({"code": code, "remplacement": remplacement, "nom_composant": nom, "quantite": int(quantite)})
                st.success("PDR enregistrée.")
                st.rerun()
            except Exception as e:
                st.error(str(e))
    delcode = st.text_input("Code à supprimer", key="pdr_delcode")
    if st.button("Supprimer PDR", key="btn_del_pdr"):
        delete_pdr_by_code(delcode.strip())
        st.success("PDR supprimée.")
        st.rerun()


# ---------------------------
# Page Export Excel
# ---------------------------
def page_export():
    st.header("Export Excel")
    bons = read_bons()
    if not bons:
        st.info("Aucun bon à exporter.")
        return
    if st.button("Générer & télécharger Excel", key="btn_gen_export"):
        try:
            excel_bytes = export_excel(bons)
            st.download_button("Télécharger bon_travail_export.xlsx", data=excel_bytes, file_name="bon_travail_export.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(str(e))

# ---------------------------
# Router - affichage des pages
# ---------------------------
if menu == "Dashboard":
    page_dashboard()
elif menu == "Production":
    page_bons("Production")
elif menu == "Maintenance":
    page_bons("Maintenance")
elif menu == "Qualité":
    page_bons("Qualité")
elif menu == "Pièces (PDR)":
    page_pdr()
elif menu == "Export Excel":
    page_export()

# Footer
st.sidebar.markdown("---")
st.sidebar.caption("⚠️ Données stockées localement sur l'instance Streamlit et sur le disque du PC poste (Othman)")

